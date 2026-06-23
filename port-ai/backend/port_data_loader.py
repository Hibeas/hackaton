"""Load PCS port data from Excel exports into SQLite cache."""

from __future__ import annotations

import json
import logging
import os
import sqlite3
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

BASE_DIR = Path(__file__).resolve().parent
DEFAULT_DATA_DIR = BASE_DIR / "data" / "port" / "raw"
PACKAGE_DATA_DIR = BASE_DIR.parent / "data"  # port-ai/data (repo-level xlsx)
DOWNLOADS_FALLBACK = Path.home() / "Downloads"
DB_PATH = BASE_DIR / "data" / "port" / "port.db"
TERMINALS_PATH = BASE_DIR / "data" / "port" / "terminals.json"

WINDOW_DAYS = int(os.environ.get("PORT_DATA_WINDOW_DAYS", "30"))
FULL_IMPORT = os.environ.get("PORT_DATA_FULL_IMPORT", "").strip() == "1"
DEMO_PORT_CALLS_LIMIT = int(os.environ.get("PORT_CALLS_DEMO_LIMIT", "150"))


def _parse_dt(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        dt = value
    else:
        text = str(value).strip()
        if not text or text.startswith("1900"):
            return None
        normalized = text.replace(" ", "T", 1) if " " in text and "T" not in text else text
        try:
            dt = datetime.fromisoformat(normalized)
        except ValueError:
            return None
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def _iso(value: datetime | None) -> str | None:
    return value.isoformat() if value else None


def resolve_data_dir() -> Path:
    configured = os.environ.get("PORT_DATA_DIR", "").strip()
    if configured:
        path = Path(configured)
        if path.is_dir():
            return path
    for candidate in (DEFAULT_DATA_DIR, PACKAGE_DATA_DIR):
        if candidate.is_dir() and any(candidate.glob("*.xlsx")):
            return candidate
    # copy-on-first-use from Downloads
    DEFAULT_DATA_DIR.mkdir(parents=True, exist_ok=True)
    patterns = ("DspShips*.xlsx", "Codeco*.xlsx", "PortCalls*.xlsx")
    copied = False
    for pattern in patterns:
        for src in DOWNLOADS_FALLBACK.glob(pattern):
            dst = DEFAULT_DATA_DIR / src.name
            if not dst.exists():
                try:
                    import shutil

                    shutil.copy2(src, dst)
                    copied = True
                    logger.info("Copied port data %s -> %s", src.name, dst)
                except OSError as exc:
                    logger.warning("Could not copy %s: %s", src, exc)
    if copied or any(DEFAULT_DATA_DIR.glob("*.xlsx")):
        return DEFAULT_DATA_DIR
    if DOWNLOADS_FALLBACK.is_dir():
        return DOWNLOADS_FALLBACK
    return DEFAULT_DATA_DIR


def load_terminals_config() -> dict[str, Any]:
    if TERMINALS_PATH.is_file():
        with TERMINALS_PATH.open(encoding="utf-8") as handle:
            return json.load(handle)
    return {"terminals": {}, "berths": {}, "terminal_aliases": {}}


def _find_files(data_dir: Path) -> dict[str, list[Path]]:
    return {
        "dsp_ships": sorted(data_dir.glob("DspShips*.xlsx")),
        "codeco": sorted(data_dir.glob("Codeco*.xlsx")),
        "port_calls": sorted(data_dir.glob("PortCalls*.xlsx")),
    }


def _init_db(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS vessels (
            imo TEXT PRIMARY KEY,
            name TEXT,
            loa_m REAL,
            dwt REAL,
            ship_type TEXT,
            mmsi TEXT,
            beam_m REAL,
            gt REAL
        );
        CREATE TABLE IF NOT EXISTS port_calls (
            call_id TEXT PRIMARY KEY,
            port_code TEXT,
            port_name TEXT,
            berth_name TEXT,
            ship_imo TEXT,
            ship_name TEXT,
            eta TEXT,
            ata TEXT,
            etd TEXT,
            atd TEXT,
            status TEXT,
            voyage_no TEXT
        );
        CREATE TABLE IF NOT EXISTS container_moves (
            move_id TEXT PRIMARY KEY,
            terminal TEXT,
            timestamp TEXT,
            full_empty TEXT,
            load_port TEXT,
            unload_port TEXT,
            direction TEXT,
            status_code TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_moves_ts ON container_moves(timestamp);
        CREATE INDEX IF NOT EXISTS idx_moves_terminal ON container_moves(terminal);
        CREATE INDEX IF NOT EXISTS idx_calls_eta ON port_calls(eta);
        """
    )


def _in_window(dt: datetime | None, window_start: datetime, window_end: datetime) -> bool:
    if dt is None:
        return False
    return window_start <= dt <= window_end


def import_excel_to_sqlite(*, force: bool = False) -> bool:
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl not installed — port Excel data unavailable")
        return False

    data_dir = resolve_data_dir()
    files = _find_files(data_dir)
    if not any(files.values()):
        logger.warning("No port Excel files found in %s", data_dir)
        return False

    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    if DB_PATH.exists() and not force:
        age_hours = (datetime.now().timestamp() - DB_PATH.stat().st_mtime) / 3600.0
        if age_hours < 6:
            return True

    now = datetime.now(timezone.utc)
    window_start = now - timedelta(days=WINDOW_DAYS)
    window_end = now + timedelta(hours=48)

    conn = sqlite3.connect(DB_PATH)
    _init_db(conn)
    conn.execute("DELETE FROM vessels")
    conn.execute("DELETE FROM port_calls")
    conn.execute("DELETE FROM container_moves")

    for path in files["dsp_ships"]:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))[2:]
        for row in rows:
            imo = str(row[7]).strip() if row[7] else None
            if not imo:
                continue
            conn.execute(
                """
                INSERT OR REPLACE INTO vessels (imo, name, loa_m, dwt, ship_type, mmsi, beam_m, gt)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    imo,
                    row[4],
                    row[3],
                    row[1],
                    row[11],
                    str(row[8]) if row[8] else None,
                    row[0],
                    row[2],
                ),
            )
        wb.close()

    port_call_candidates: list[tuple[datetime | None, tuple[Any, ...]]] = []
    for path in files["port_calls"]:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))[2:]
        for row in rows:
            eta = _parse_dt(row[32])
            ata = _parse_dt(row[30])
            etd = _parse_dt(row[33])
            atd = _parse_dt(row[31])
            call_id = str(row[36] or f"{row[21]}-{row[32]}")
            status = "in_port" if ata and not atd else "expected" if eta and not ata else "departed"
            sort_dt = ata or eta or atd or etd
            if FULL_IMPORT or any(
                _in_window(dt, window_start, window_end) for dt in (eta, ata, etd, atd)
            ):
                port_call_candidates.append(
                    (
                        sort_dt,
                        (
                            call_id,
                            row[11],
                            row[13],
                            row[2],
                            str(row[21]) if row[21] else None,
                            row[20],
                            _iso(eta),
                            _iso(ata),
                            _iso(etd),
                            _iso(atd),
                            status,
                            str(row[35]) if row[35] else None,
                        ),
                    )
                )
        wb.close()

    if not port_call_candidates and files["port_calls"]:
        for path in files["port_calls"]:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))[2:]
            for row in rows:
                eta = _parse_dt(row[32])
                ata = _parse_dt(row[30])
                atd = _parse_dt(row[31])
                etd = _parse_dt(row[33])
                call_id = str(row[36] or f"{row[21]}-{row[32]}")
                status = "in_port" if ata and not atd else "expected" if eta and not ata else "departed"
                port_call_candidates.append(
                    (
                        eta or ata or atd or etd,
                        (
                            call_id,
                            row[11],
                            row[13],
                            row[2],
                            str(row[21]) if row[21] else None,
                            row[20],
                            _iso(eta),
                            _iso(ata),
                            _iso(etd),
                            _iso(atd),
                            status,
                            str(row[35]) if row[35] else None,
                        ),
                    )
                )
            wb.close()

    port_call_candidates.sort(key=lambda item: item[0] or datetime.min.replace(tzinfo=timezone.utc), reverse=True)
    limit = len(port_call_candidates) if FULL_IMPORT else min(DEMO_PORT_CALLS_LIMIT, len(port_call_candidates))
    for _, payload in port_call_candidates[:limit]:
        conn.execute(
            """
            INSERT OR REPLACE INTO port_calls
            (call_id, port_code, port_name, berth_name, ship_imo, ship_name, eta, ata, etd, atd, status, voyage_no)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            payload,
        )

    move_id = 0
    for path in files["codeco"]:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))[2:]
        for row in rows:
            ts = _parse_dt(row[2])
            if not FULL_IMPORT and not _in_window(ts, window_start, now):
                continue
            terminal = row[1] or row[0]
            if not terminal:
                continue
            load_port = row[6]
            unload_port = row[5]
            direction = "export" if load_port and str(load_port).startswith("PL") else "import"
            move_id += 1
            conn.execute(
                """
                INSERT OR REPLACE INTO container_moves
                (move_id, terminal, timestamp, full_empty, load_port, unload_port, direction, status_code)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    f"{path.stem}-{move_id}",
                    str(terminal),
                    _iso(ts),
                    row[4],
                    load_port,
                    unload_port,
                    direction,
                    row[3],
                ),
            )
        wb.close()

    conn.commit()
    conn.close()
    logger.info(
        "Port SQLite import complete: vessels + calls + moves in %s (window=%sd full=%s)",
        DB_PATH,
        WINDOW_DAYS,
        FULL_IMPORT,
    )
    return True


class PortDataStore:
    def __init__(self) -> None:
        self._loaded = False
        self._vessels: dict[str, dict[str, Any]] = {}
        self._port_calls: list[dict[str, Any]] = []
        self._container_moves: list[dict[str, Any]] = []
        self._terminals = load_terminals_config()
        self.updated_at: datetime | None = None

    def refresh(self, *, force_import: bool = False) -> None:
        if import_excel_to_sqlite(force=force_import) and DB_PATH.is_file():
            self._load_from_db()
        self._loaded = True
        self.updated_at = datetime.now(timezone.utc)

    def _load_from_db(self) -> None:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        self._vessels = {
            row["imo"]: dict(row)
            for row in conn.execute("SELECT * FROM vessels")
        }
        self._port_calls = [dict(row) for row in conn.execute("SELECT * FROM port_calls")]
        self._container_moves = [dict(row) for row in conn.execute("SELECT * FROM container_moves")]
        conn.close()

    @property
    def vessels(self) -> dict[str, dict[str, Any]]:
        return self._vessels

    @property
    def port_calls(self) -> list[dict[str, Any]]:
        return self._port_calls

    @property
    def container_moves(self) -> list[dict[str, Any]]:
        return self._container_moves

    @property
    def terminals_config(self) -> dict[str, Any]:
        return self._terminals

    def vessel_for_imo(self, imo: str | None) -> dict[str, Any] | None:
        if not imo:
            return None
        return self._vessels.get(str(imo))

    def summary(self) -> dict[str, Any]:
        return {
            "vessel_count": len(self._vessels),
            "port_call_count": len(self._port_calls),
            "container_move_count": len(self._container_moves),
            "updated_at": self.updated_at.isoformat() if self.updated_at else None,
            "data_dir": str(resolve_data_dir()),
            "db_path": str(DB_PATH),
        }


port_data_store = PortDataStore()
